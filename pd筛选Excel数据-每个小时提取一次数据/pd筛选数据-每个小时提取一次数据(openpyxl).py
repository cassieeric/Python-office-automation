# -*- coding: utf-8 -*-
# @Time : 2022/4/21 18:00
# @Author : Python进阶者
# 公众号：Python爬虫与数据挖掘
# @FileName : pd筛选数据-每个小时提取一次数据(openpyxl).py

from openpyxl import load_workbook, Workbook
from datetime import datetime

# 打开数据工作簿
workbook = load_workbook('数据.xlsx')
# 打开工作表
sheet = workbook.active
time_column = sheet['C']
row_lst = []
date_lst = []
hour_lst = []
for cell in time_column:
    if cell.value != "SampleTime" and cell.value != None:
        # print(cell.value.date())
        if cell.value.date() not in date_lst:
            date_lst.append(cell.value.date())
        # row_lst.append(cell.row)
print(date_lst)

# if all(cell.value != "SampleTime", cell.value != None, cell.value.date() == date, cell.value.hour not in hour_lst):

for date in date_lst:
    # print(date)
    for cell in time_column:
        # if all((cell.value != "SampleTime", cell.value != None, cell.value.date() == date, cell.value.hour not in hour_lst)):
        #     row_lst.append(cell.row)
        if cell.value != "SampleTime" and cell.value != None:
            if cell.value.date() == date:
                if cell.value.hour not in hour_lst:
                    hour_lst.append(cell.value.hour)
                    row_lst.append(cell.row)
    hour_lst = []
print(hour_lst)
# 将满足要求的数据写入到新表
new_workbook = Workbook()
new_sheet = new_workbook.active

# 创建和原数据 一样的表头（第一行）
header = sheet[1]
header_lst = []
for cell in header:
    header_lst.append(cell.value)
new_sheet.append(header_lst)

# 从旧表中根据行号提取符合条件的行，并遍历单元格获取值，以列表形式写入新表
for row in row_lst:
    data_lst = []
    for cell in sheet[row]:
        data_lst.append(cell.value)
    new_sheet.append(data_lst)

# 最后切记保存
new_workbook.save('新表.xlsx')
print("满足条件的新表保存完成！")
