#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd

src_dir = "."
supplier = "xxx"
df = pd.read_excel("原始数据.xlsx")
df


# In[2]:
from glob import glob

from win32com import client as win32
import os

word = win32.Dispatch("Word.Application")
word.Visible = True
word.ScreenUpdating = False
word_data = {}
files = glob(f"{src_dir}/物资合同*.doc*")
print("加载word数据")
first = None
for file in files:
    filepath = os.path.abspath(file)
    print(filepath)
    rDoc = word.Documents.Open(filepath)
    header, code = None, None
    data = []
    for row in rDoc.Tables(1).Rows:
        row = [c.Range.Text.strip("\r\x07") for c in row.Cells]
        if header is None:
            if row[0] == "序号":
                header = row
            elif len(row) > 2 and row[1].startswith("xxx编号"):
                code = row[1].split("：")[1]
        else:
            if row[0] == "合计金额：":
                break
            data.append(row)
    data = pd.DataFrame(data, columns=header)
    print(code)
    data["供应商"] = supplier
    data.采购数量 = pd.to_numeric(data.采购数量)
    data.价款小计 = pd.to_numeric(data.价款小计)
    word_data[code] = data.iloc[:, [2, 1, 12, 5, 7, 8]]
#     display(word_data[code])
    if first is None:
        first = rDoc
    else:
        rDoc.Close()
word.ScreenUpdating = True
first.Close()
word.Quit()


# In[3]:


from copy import copy
from openpyxl import load_workbook
import re


for i, s in df.iloc[0:].iterrows():
    print(s.values)
    code = s["编号"]
    c2 = s["编号"]
    city = s["订单城市"]
    wdf = word_data.get(code)
    if wdf is None:
        print(code, "对应合同编号的Word文档未找到")
        continue
    wb = load_workbook("模板.xlsx")
    sht = wb.active
    sht["B3"].value = s["编号"]
    sht["E3"].value = f"{s['金额（不含税）']:.2f}元"
    sht["A8"].value = code
    sht["E8"].value = s["合同金额（不含税，元）"]
    sht["E10"].value = s["合同金额（不含税，元）"]

    height = sht.row_dimensions[12].height
    styles = []
    for cell, in sht.iter_cols(1, 6, 12, 12):
        tmp = {}
        tmp["border"] = copy(cell.border)
        tmp["fill"] = copy(cell.fill)
        tmp["font"] = copy(cell.font)
        tmp["alignment"] = copy(cell.alignment)
        tmp["number_format"] = copy(cell.number_format)
        styles.append(tmp)
    sht.delete_rows(12)
    n = wdf.shape[0]
    sht.insert_rows(12, n)
    for i, row in enumerate(sht.iter_rows(12, 11+n, max_col=6)):
        sht.row_dimensions[12+i].height = height
        for j, cell in enumerate(row):
            cell.value = wdf.iat[i, j]
            cell.border = styles[j]["border"]
            cell.fill = styles[j]["fill"]
            cell.font = styles[j]["font"]
            cell.number_format = styles[j]["number_format"]
            cell.alignment = styles[j]["alignment"]
    sht[f"E{n+12}"].value = f"=sum(E12:E{n+11})"
    sht[f"F{n+12}"].value = f"=sum(F12:F{n+11})"
    text = wdf["物料"].str.replace(
        "-(?:-)?|-?\d.*", "", regex=True)
    text = (wdf.采购数量.astype(str)+wdf.计量单位+text).str.cat(sep="、")
    nt = "零一二三四五六七八九十"
    addr, nums = re.split("(?=\d)", city, maxsplit=1)
    t = []
    for num in nums:
        t.append(nt[int(num)])
    num = "".join(t)
    text = f"{text}"
    sht["B6"].value = text
    wb.save(f"xxx-{code}.xlsx")
    print("保存到", f"xxx-{code}.xlsx")


# In[4]:


xlApp = win32.Dispatch("Excel.Application")
xlApp.Visible = True
xlApp.ScreenUpdating = False
xlApp.DisplayAlerts = False

files = glob(f"{src_dir}/xxx-*.xlsx")
print("加载excel结果数据")
first = None
try:
    for file in files:
        filepath = os.path.abspath(file)
        print(filepath)
        book = xlApp.Workbooks.Open(filepath, ReadOnly=1)
        book.ExportAsFixedFormat(0, filepath[:-4]+"pdf")
        print("保存到", filepath[:-4]+"pdf")
        book.Save()
        book.Close()
finally:
    xlApp.ScreenUpdating = True
    xlApp.Quit()


# In[ ]:
