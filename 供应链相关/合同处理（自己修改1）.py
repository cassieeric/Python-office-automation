#!/usr/bin/env python
# coding: utf-8

from glob import glob

from win32com import client as win32
import os
import pandas as pd

src_dir = "."
supplier = "深圳市佳贤通信设备有限公司"
df = pd.read_excel("小站订单20230314.xlsx", sheet_name='佳贤')

word = win32.Dispatch("Word.Application")
word.Visible = True
word.ScreenUpdating = False
word_data = {}
files = glob(f"{src_dir}/物资合同*.doc*")
print("加载word数据")
first = None
for file in files:
    filepath = os.path.abspath(file)
    print(filepath)
    rDoc = word.Documents.Open(filepath)
    header, code = None, None
    data = []
    for row in rDoc.Tables(1).Rows:
        row = [c.Range.Text.strip("\r\x07") for c in row.Cells]
        if header is None:
            if row[0] == "序号":
                header = row
            elif len(row) > 2 and row[1].startswith("发货通知单编号"):
                code = row[1].split("：")[1]
        else:
            if row[0] == "货物费合计金额：":
                break
            data.append(row)
    data = pd.DataFrame(data, columns=header)
    print(code)
    data["供应商"] = supplier
    data.采购数量 = pd.to_numeric(data.采购数量)
    data.价款小计 = pd.to_numeric(data.价款小计)
    word_data[code] = data.iloc[:, [2, 1, 12, 5, 7, 8]]
#     display(word_data[code])
    if first is None:
        first = rDoc
    else:
        rDoc.Close()
word.ScreenUpdating = True
first.Close()
word.Quit()


# In[3]:
from copy import copy
from openpyxl import load_workbook
import re

# for i, s in df.iterrows():
# for i, s in df.iloc[218:].iterrows():
for i, s in df.iloc[217:218].iterrows():  # 这个序号要比表格中的index要少2，比方说江苏南通5的index是219，那么这里应该写为217:218
    print("正在处理：", s.values[2])
    code = s["订单编号"]
    c2 = s["销售订单编号"]
    city = s["城市"]
    wdf = word_data.get(code)
    if wdf is None:
        print(code, "对应合同编号的Word文档未找到")
        continue
    wb = load_workbook("佳贤物资出库申请单模板.xlsx")
    sht = wb.active
    sht["B3"].value = s["销售订单编号"]
    # sht["E3"].value = str(s["销售订单设备总价（不含税）"])+"元"
    # sht["E3"].value = str(float(s["销售订单设备总价（不含税）"])) + "元"
    sht["E3"].value = f"{s['销售订单设备总价（不含税）']:.2f}元"
    sht["A8"].value = code
    sht["E8"].value = s["合同金额（不含税）"]
    sht["E10"].value = s["合同金额（不含税）"]

    border = copy(sht["A12"].border)
    fill = copy(sht["A12"].fill)
    font = copy(sht["A12"].font)
    alignment = copy(sht["A12"].alignment)
    height = sht.row_dimensions[12].height
    number_format = "0.00_ "
    # number_format = copy(sht["F12"].number_format)
    sht.delete_rows(12)
    n = wdf.shape[0]
    sht.insert_rows(12, n)
    for i, row in enumerate(sht.iter_rows(12, 11+n, max_col=6)):
        sht.row_dimensions[12+i].height = height
        for j, cell in enumerate(row):
            cell.value = wdf.iat[i, j]
            cell.border = border
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            if j == 5:
                cell.number_format = number_format
    sht[f"E{n+12}"].value = f"=sum(E12:E{n+11})"
    # sht[f"F{n+12}"].value = f"=sum(F12:F{n+11})"
    sht[f"F{n+12}"].value = f"=sum(F12:F{n+11})"
    text = wdf["货物名称（物料名称）"].str.replace(
        "5G小基站-(?:自研EXT型-)?|-?\d.*", "", regex=True)
    text = (wdf.采购数量.astype(str)+wdf.计量单位+text).str.cat(sep="、")
    nt = "零一二三四五六七八九十"
    addr, nums = re.split("(?=\d)", city, maxsplit=1)
    t = []
    for num in nums:
        t.append(nt[int(num)])
    num = "".join(t)
    text = f"领用后，部件组装后对应{text}。上电、调试，测试合格后，按照合同号：{c2}发货到{addr}第{num}批。"
    sht["B6"].value = text
    wb.save(f"物资出库申请-{code}.xlsx")
    print("保存到", f"物资出库申请-{code}.xlsx")


# In[5]:


xlApp = win32.Dispatch("Excel.Application")
xlApp.Visible = True
xlApp.ScreenUpdating = False
xlApp.DisplayAlerts = False

files = glob(f"{src_dir}/物资出库申请-*.xlsx")
print("加载excel结果数据")
first = None
try:
    for file in files:
        filepath = os.path.abspath(file)
        print(filepath)
        book = xlApp.Workbooks.Open(filepath, ReadOnly=1)
        book.ExportAsFixedFormat(0, filepath[:-4]+"pdf")
        print("保存到", filepath[:-4]+"pdf")
        book.Save()
        book.Close()
finally:
    xlApp.ScreenUpdating = True
    xlApp.Quit()


# In[ ]:




