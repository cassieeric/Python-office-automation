#!/usr/bin/env python
# coding: utf-8

# In[1]:
import pandas as pd

global hetong_number
global code
# global text
# global addr
# global num
import time
from docx2pdf import convert
from win32com.client import Dispatch
from copy import copy
from openpyxl import load_workbook
import re


src_dir = r"./待处理的物资合同"
supplier = "深圳市佳贤通信科技股份有限公司"
df = pd.read_excel("小站订单20230323.xlsx", sheet_name='佳贤')


# In[2]:


from glob import glob
from win32com import client as win32
import os

word = win32.Dispatch("Word.Application")
word.Visible = True
word.ScreenUpdating = False
word_data = {}
files = glob(f"{src_dir}/物资合同*.doc*")
print("加载word数据")
first = None
for file in files:
    filepath = os.path.abspath(file)
    print(filepath)
    rDoc = word.Documents.Open(filepath)
    header, code = None, None
    data = []
    for row in rDoc.Tables(1).Rows:
        row = [c.Range.Text.strip("\r\x07") for c in row.Cells]
        if header is None:
            if row[0] == "序号":
                header = row
            elif len(row) > 2 and row[1].startswith("发货通知单编号"):
                code = row[1].split("：")[1]
        else:
            if row[0] == "货物费合计金额：":
                break
            data.append(row)
    data = pd.DataFrame(data, columns=header)
    print(code)
    data["供应商"] = supplier
    data.采购数量 = pd.to_numeric(data.采购数量)
    data.价款小计 = pd.to_numeric(data.价款小计)
    word_data[code] = data.iloc[:, [2, 1, 12, 5, 7, 8]]
#     display(word_data[code])
    if first is None:
        first = rDoc
    else:
        rDoc.Close()
word.ScreenUpdating = True
first.Close()
word.Quit()


def convert_word_2_pdf(filename):
    print(filename)
    word_file = fr'C:\Users\pdcfi\Desktop\出库申请单制作\佳贤出库申请单制作\{filename}'
    word = Dispatch('Word.Application')
    # 以Doc对象打开文件
    doc_ = word.Documents.Open(word_file)
    # 另存为pdf文件
    doc_.SaveAs(word_file.replace(".docx", ".pdf"), FileFormat=17)
    # 关闭doc对象
    doc_.Close()
    # 退出word对象
    word.Quit()
    print("物资开箱报告word格式转pdf格式已完成！")
# In[3]:


def kaixiangbaogao(text, addr, nums):
    # In[5 ]: 制作物资开箱报告
    # 附件2：物资开箱验收报告（模板0723）（HQDD202302240016）.docx
    # 物资开箱验收报告模板.docx
    import docx

    template = "物资开箱验收报告模板.docx"  # 模板的文档名
    hetong_placeholder = "<合同编号>"  # 需要被替换的字符，也被称为占位符
    dingdan_placeholder = "placeholder"  # 需要被替换的字符，也被称为占位符
    reason_placeholder = "<领用原因>"  # 需要被替换的字符，也被称为占位符
    doc = docx.Document(template)
    print(f"正在对物资开箱验收报告内容替换..，.")
    # 循环遍历每个段落
    for para in doc.paragraphs:
        # 循环遍历段落中的每一组文字
        for run in para.runs:
            if hetong_placeholder in run.text:
                # 用具体的名字替换占位符
                run.text = run.text.replace(hetong_placeholder, str(hetong_number))

            if dingdan_placeholder in run.text:
                run.text = run.text.replace(dingdan_placeholder, code)

            if reason_placeholder in run.text:
                lingyong_reason = f"此订单对应的设备发货到{addr}第{nums}批，部件组装后对应{text}。设备及配置清单请见订单以及到货证明附件。"
                run.text = run.text.replace(reason_placeholder, lingyong_reason)
        filename = f'附件2：物资开箱验收报告（模板0723）（{code}）.docx'
        doc.save(filename)
        print(f"物资开箱验收报告内容替换完成！")

        # print(f"开始对物资开箱验收报告word版本进行转pdf...")
        # time.sleep(3)  # 不睡眠的话，就报错：Word.Application.Documents，可能是上一个pdf没存完
        # convert_word_2_pdf(filename)  # word保存之后，同时转换成pdf文档
        # print(f"物资开箱验收报告word版本转pdf已经完成！")


for i, s in df.iloc[231:239].iterrows():  # 这个序号要比表格中的index要少2，比方说江苏南通5的index是219，那么这里应该写为217:218
    print("正在处理：", s.values[2])
    code = s["订单编号"]
    c2 = s["销售订单编号"]
    city = s["城市"]
    wdf = word_data.get(code)
    if wdf is None:
        print(code, "对应合同编号的Word文档未找到")
        continue
    wb = load_workbook("佳贤物资出库申请单模板.xlsx")
    sht = wb.active
    sht["B3"].value = s["销售订单编号"]
    sht["E3"].value = f"{s['销售订单设备总价（不含税）']:.2f}元"
    sht["A8"].value = code
    sht["E8"].value = s["合同金额（不含税）"]
    sht["E10"].value = s["合同金额（不含税）"]
    hetong_number = s["合同编号"]

    height = sht.row_dimensions[12].height
    styles = []
    for cell, in sht.iter_cols(1, 6, 12, 12):
        tmp = {}
        tmp["border"] = copy(cell.border)
        tmp["fill"] = copy(cell.fill)
        tmp["font"] = copy(cell.font)
        tmp["alignment"] = copy(cell.alignment)
        tmp["number_format"] = copy(cell.number_format)
        styles.append(tmp)
    sht.delete_rows(12)
    n = wdf.shape[0]
    sht.insert_rows(12, n)
    for i, row in enumerate(sht.iter_rows(12, 11 + n, max_col=6)):
        sht.row_dimensions[12 + i].height = height
        for j, cell in enumerate(row):
            cell.value = wdf.iat[i, j]
            cell.border = styles[j]["border"]
            cell.fill = styles[j]["fill"]
            cell.font = styles[j]["font"]
            cell.number_format = styles[j]["number_format"]
            cell.alignment = styles[j]["alignment"]
    sht[f"E{n + 12}"].value = f"=sum(E12:E{n + 11})"
    sht[f"F{n + 12}"].value = f"=sum(F12:F{n + 11})"
    t = pd.DataFrame()
    a = wdf["货物名称（物料名称）"]
    t[["num", "unit"]] = a.str.extract("-(\d+)m*(m|m2)-", expand=False)
    t.unit = t.unit.map({"m": "米", "m2": "平方"})
    t.fillna("", inplace=True)
    text = wdf["货物名称（物料名称）"].str.replace(
        "5G小基站-(?:自研EXT型-)?|-?\d.*", "", regex=True)
    text = (wdf.采购数量.astype(str) + wdf.计量单位 + t.num + t.unit + text).str.cat(sep="、")
    # nt = "零一二三四五六七八九十"
    addr, nums = re.split("(?=\d)", city, maxsplit=1)
    # t = []
    # for num in nums:
    #     t.append(nt[int(num)])
    # num = "".join(t)

    final_text = f"领用后，部件组装后对应{text}。上电、调试，测试合格后，按照合同号：{c2}发货到{addr}第{nums}批。"
    sht["B6"].value = final_text
    wb.save(f"物资出库申请-{code}.xlsx")
    print("保存到: ", f"物资出库申请-{code}.xlsx")
    kaixiangbaogao(text, addr, nums)


# In[4]: Excel格式转pdf
xlApp = win32.Dispatch("Excel.Application")
xlApp.Visible = True
xlApp.ScreenUpdating = False
xlApp.DisplayAlerts = False

files = glob(f"./物资出库申请-*.xlsx")
print("加载excel结果数据")
first = None
try:
    for file in files:
        filepath = os.path.abspath(file)
        print(filepath)
        book = xlApp.Workbooks.Open(filepath, ReadOnly=1)
        book.ExportAsFixedFormat(0, filepath[:-4]+"pdf")
        print("保存到", filepath[:-4]+"pdf")
        book.Save()
        book.Close()
finally:
    xlApp.ScreenUpdating = True
    xlApp.Quit()

# In[6 ]: 制作物资开箱报告word转pdf
# word文档处理器
# from win32com.client import Dispatch

# 文件目录遍历器
# from os import walk
# import time
# time.sleep(3)  # 不睡眠的话，就报错：Word.Application.Documents，可能是上一个pdf没存完

# 方法一
# from docx2pdf import convert
# convert(f'./附件2：物资开箱验收报告（模板0723）（{code}）.docx', f'./附件2：物资开箱验收报告（模板0723）（{code}）.pdf')
# print("物资开箱报告word格式转pdf格式已完成！")

# 方法二
# doc_path = r'C:\Users\pdcfi\Desktop\出库申请单制作\佳贤出库申请单制作'
# for root, dirs, filenames in walk(doc_path):
#     # 遍历当前文件名称、并校验是否是word文档
#     for file in filenames:
#         # if file.endswith(".doc") or file.endswith(".docx"):
#         if file.startswith("附件2") and file.endswith(".docx"):
#             word_file = str(root + "\\" + file)
#             # 如果当前文件是word文档则调用word转换函数
#             word = Dispatch('Word.Application')
#             # 以Doc对象打开文件
#             doc_ = word.Documents.Open(word_file)
#             # 另存为pdf文件
#             doc_.SaveAs(word_file.replace(".docx", ".pdf"), FileFormat=17)
#             # 关闭doc对象
#             doc_.Close()
#             # 退出word对象
#             word.Quit()
# print("物资开箱报告word格式转pdf格式已完成！")
