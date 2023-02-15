import openpyxl
import time

workbook1 = openpyxl.load_workbook("上市公司收入确认表(研究院)_模板.xlsx")
worksheet1 = workbook1.worksheets[0]

workbook2 = openpyxl.load_workbook("小站订单.xlsx")
worksheet2 = workbook2['省对研究院']
# print(worksheet2.max_row)
for i in range(98):
    i = i + 3
    print(f"正在第{i}行，处理订单：{worksheet2[f'C{i}'].value}...")
    worksheet1['C4'].value = worksheet2[f'CU{i}'].value
    worksheet1['D4'].value = f"中国电信股份有限公司{worksheet2[f'DM{i}'].value}分公司"
    worksheet1['F4'].value = worksheet2[f'D{i}'].value
    new_file_name = f"上市公司收入确认表(研究院)  （{worksheet2[f'C{i}'].value} {worksheet2[f'D{i}'].value}）"
    workbook1.save(f'./生成的收入确认表/{new_file_name}' + '.xlsx')
    time.sleep(3)
    print(f"订单：{worksheet2[f'C{i}'].value}处理完成")
